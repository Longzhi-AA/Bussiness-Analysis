import pandas as pd
from openpyxl import load_workbook


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]
    Returns: None
    """
    # from openpyxl import load_workbook

    # import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', date_format='MM/DD/YYYY', datetime_format='MM/DD/YYYY')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, header=False, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


# if __name__ == '__main__':
# #     xl_path = 'E:/BI/DBS reconc/dataset_inventory_gap_analysis.xlsx'
# #     snap_path = 'E:/BI/DBS reconc/data_snapshot date_06.05.xlsx'
# #
# #     datafildes = ['MFG Part#', 'Region', 'OH Qty', 'Cluster', 'Program.1', 'Location', 'IPN', 'SysCost($)',
# #                   'Last Refresh Date']
# #     datafildes2 = ['MFG Part#', 'Region', 'OH Qty', 'Cluster', 'Program', 'Location', 'IPN', 'SysCost($)',
# #                    'Last Refresh Date']
# #     hyve_df = pd.read_excel(snap_path, sheet_name='Hyve OH INV')
# #     hyve_df = hyve_df[datafildes]
# #     hyve_df.columns = datafildes2
# #     dbs_df = pd.read_excel(snap_path, sheet_name='DBS OH INV')
# #     dbs_df = dbs_df[datafildes2]
# #     print('concat new dateframe...')
# #     oh_df = pd.concat([hyve_df, dbs_df], axis=0, ignore_index=True, sort=False)
# #     oh_df['OH Qty'].astype('int')
# #     oh_df['SysCost($)'].astype('float')
# #     total_oh_df = pd.read_excel(xl_path, sheet_name='Total_OH_INV')
# #     rows = total_oh_df['MFG Part#'].count()
# #     print('append data to excel...')
# #     try:
# #         append_df_to_excel(xl_path, oh_df, sheet_name='Total_OH_INV', index=False, startrow=rows + 1, startcol=0)
# #         print('****Successfully append data to excel!')
# #     except Exception as e:
# #         print('****Something wrong when appending data, error message:{}'.format(e))
